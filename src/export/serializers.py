"""Módulo de serialización de datos a CSV y Excel (XLSX) con seguridad anti-inyección."""

import csv
import re
from io import BytesIO, StringIO
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
_CTRL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitizar_formula(val: object) -> Any:
    """Sanitiza inyecciones de fórmulas y elimina caracteres de control."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return val
    s_val = _CTRL_CHARS.sub("", str(val))
    if s_val.startswith(_FORMULA_PREFIXES):
        return "'" + s_val
    return s_val


def a_csv(columnas: list[str], filas: list[dict[str, object]]) -> bytes:
    """Serializa columnas y filas a formato CSV RFC 4180 con BOM UTF-8 y anti-inyección."""
    output = StringIO()
    writer = csv.writer(output, lineterminator="\r\n")

    # Escribir encabezados sanitizados
    encabezados = [_sanitizar_formula(col) for col in columnas]
    writer.writerow(encabezados)

    # Escribir filas sanitizadas
    for fila in filas:
        row_data: list[object] = []
        for col in columnas:
            v = fila.get(col)
            row_data.append(_sanitizar_formula(v))
        writer.writerow(row_data)

    # BOM UTF-8 (\xef\xbb\xbf) + contenido codificado en utf-8
    csv_str = output.getvalue()
    return b"\xef\xbb\xbf" + csv_str.encode("utf-8")


def _escribir_celda_xlsx(
    cell: Any,
    raw_val: object,
    row_fill: PatternFill | None,
    font: Font,
    border: Border,
    alignment: Alignment,
) -> None:
    """Escribe el valor sanitizado y aplica los estilos visuales a la celda XLSX."""
    cell.value = _sanitizar_formula(raw_val)

    if row_fill:
        cell.fill = row_fill
    cell.font = font
    cell.border = border
    cell.alignment = alignment


def a_xlsx(columnas: list[str], filas: list[dict[str, object]]) -> bytes:
    """Serializa columnas y filas a formato XLSX (openpyxl) estilizado y seguro contra fórmulas."""
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Report"

    # Estilos corporativos
    header_fill = PatternFill(start_color="A5C8ED", end_color="A5C8ED", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

    header_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    cell_font = Font(name="Segoe UI", size=10, color="000000")

    thin_border_side = Side(border_style="thin", color="CCCCCC")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align_center = Alignment(horizontal="center", vertical="center")
    cell_align_left = Alignment(horizontal="left", vertical="center")

    # 1. Escribir encabezados
    for col_idx, col_name in enumerate(columnas, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.value = _sanitizar_formula(col_name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

    ws.row_dimensions[1].height = 28

    # 2. Escribir filas de datos
    for row_idx, fila in enumerate(filas, start=2):
        row_fill = zebra_fill if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(columnas, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            raw_val = fila.get(col_name)
            align = cell_align_left if col_idx == 1 else cell_align_center
            _escribir_celda_xlsx(c, raw_val, row_fill, cell_font, thin_border, align)

        ws.row_dimensions[row_idx].height = 20

    # 3. Auto-ajuste de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_col_idx = col[0].column or 1
        col_letter = get_column_letter(col_col_idx)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
