"""Pruebas del endpoint POST /v1/exportar."""

from io import BytesIO
from typing import Any
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from src.api.main import app
from src.sql.executor import ResultadoEjecucion

client = TestClient(app)


def _ok(**kw: Any) -> ResultadoEjecucion:
    base: dict[str, Any] = {
        "ok": True,
        "columnas": ["a", "b"],
        "filas": [{"a": 1, "b": "x"}],
        "n_filas": 1,
    }
    base.update(kw)
    return ResultadoEjecucion(**base)


def test_export_csv_ok() -> None:
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=_ok()),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a,b FROM ventas", "formato": "csv"})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    assert 'attachment; filename="export_' in r.headers["content-disposition"]
    assert r.content.startswith(b"\xef\xbb\xbf")


def test_export_xlsx_ok() -> None:
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=_ok()),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a,b FROM ventas", "formato": "xlsx"})
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.content)).active
    assert ws is not None
    assert ws.cell(row=1, column=1).value == "a"


def test_export_antifuga_pyodbc() -> None:
    fuga = ResultadoEjecucion(
        ok=False,
        codigo_error="EJECUCION_SQL",
        mensaje="[Microsoft][ODBC Driver 18] Invalid column 'xyz' en la tabla ventas",
    )
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=fuga),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT xyz FROM ventas", "formato": "csv"})
    assert r.status_code == 400
    assert "[Microsoft][ODBC" not in r.text
    assert "xyz" not in r.text


def test_export_infra_bd_es_503_sin_fuga() -> None:
    infra = ResultadoEjecucion(
        ok=False, codigo_error="EJECUCION_BD", mensaje="conn string pwd=secreto"
    )
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=infra),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a FROM ventas", "formato": "csv"})
    assert r.status_code == 503
    assert "pwd" not in r.text


def test_export_inyeccion_formula_end_to_end() -> None:
    peligro = _ok(columnas=["a"], filas=[{"a": "=SUM(A1:A9)"}])
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=peligro),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a FROM ventas", "formato": "csv"})
    assert r.status_code == 200
    assert "'=SUM(A1:A9)" in r.content.decode("utf-8-sig")


def test_export_drop_rechazado_sin_archivo() -> None:
    """SELECT-only en la ruta de export: el validador rechaza DROP antes de tocar la BD."""
    with patch("src.api.export.esquema_columnas", return_value={"ventas": {"id"}}):
        r = client.post("/v1/exportar", json={"sql": "DROP TABLE ventas", "formato": "csv"})
    assert r.status_code == 400
    assert "text/csv" not in r.headers.get("content-type", "")


def test_export_body_sin_sql_422() -> None:
    assert client.post("/v1/exportar", json={"formato": "csv"}).status_code == 422


def test_export_formato_invalido_422() -> None:
    r = client.post("/v1/exportar", json={"sql": "SELECT a FROM ventas", "formato": "pdf"})
    assert r.status_code == 422


def test_export_error_interno_es_500() -> None:
    """un error interno del servidor debe ser 500, no 400 (culpa del cliente)."""
    interno = ResultadoEjecucion(ok=False, codigo_error="ERROR_INTERNO", mensaje="boom interno")
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=interno),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a FROM ventas", "formato": "csv"})
    assert r.status_code == 500


def test_export_dominio_no_ecoa_nombres_de_esquema() -> None:
    """los mensajes de error de dominio no ecoan nombres reales de tabla/columna."""
    dom = ResultadoEjecucion(
        ok=False,
        codigo_error="TABLA_O_COLUMNA_DESCONOCIDA",
        mensaje="Columna 'salario_secreto' desconocida en tabla 'nomina_confidencial'",
    )
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=dom),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT x FROM y", "formato": "csv"})
    assert r.status_code == 400
    assert "nomina_confidencial" not in r.text
    assert "salario_secreto" not in r.text


def test_export_header_truncado_presente() -> None:
    """n_filas >= export_max_rows → header X-Export-Truncated."""
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=_ok(n_filas=50000)),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a FROM ventas", "formato": "csv"})
    assert r.headers.get("x-export-truncated") == "true"


def test_export_header_truncado_ausente_en_normal() -> None:
    with (
        patch("src.api.export.esquema_columnas", return_value={}),
        patch("src.api.export.ejecutar_consulta", return_value=_ok(n_filas=5)),
    ):
        r = client.post("/v1/exportar", json={"sql": "SELECT a FROM ventas", "formato": "csv"})
    assert "x-export-truncated" not in {k.lower() for k in r.headers}
