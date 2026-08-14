"""Pruebas de los serializadores de export."""

from io import BytesIO

import pytest
from openpyxl import load_workbook
from src.export.serializers import a_csv, a_xlsx


def _csv_text(cols: list[str], filas: list[dict[str, object]]) -> str:
    return a_csv(cols, filas).decode("utf-8-sig")


def test_csv_bom_presente() -> None:
    assert a_csv(["a"], [{"a": 1}]).startswith(b"\xef\xbb\xbf")


def test_csv_encabezado_y_valores() -> None:
    txt = _csv_text(["a", "b"], [{"a": 1, "b": "x"}])
    assert "a,b" in txt
    assert "1,x" in txt


def test_csv_campo_con_coma_va_citado() -> None:
    assert '"x,y"' in _csv_text(["a"], [{"a": "x,y"}])


def test_csv_comilla_doble_se_duplica() -> None:
    assert '"di ""hola"""' in _csv_text(["a"], [{"a": 'di "hola"'}])


def test_csv_unicode() -> None:
    assert "Bogotá ñ" in _csv_text(["c"], [{"c": "Bogotá ñ"}])


def test_csv_none_es_vacio() -> None:
    assert ",1" in _csv_text(["a", "b"], [{"a": None, "b": 1}])


@pytest.mark.parametrize("prefijo", ["=", "+", "-", "@"])
def test_csv_anti_formula(prefijo: str) -> None:
    assert f"'{prefijo}SUM(A1)" in _csv_text(["a"], [{"a": f"{prefijo}SUM(A1)"}])


def test_xlsx_abre_y_encabezados() -> None:
    ws = load_workbook(BytesIO(a_xlsx(["a", "b"], [{"a": 1, "b": "x"}]))).active
    assert ws is not None
    assert ws.cell(row=1, column=1).value == "a"
    assert ws.cell(row=1, column=2).value == "b"


def test_xlsx_valor_numerico_se_preserva() -> None:
    ws = load_workbook(BytesIO(a_xlsx(["n"], [{"n": 42}]))).active
    assert ws is not None
    assert ws.cell(row=2, column=1).value == 42


def test_xlsx_anti_formula() -> None:
    ws = load_workbook(BytesIO(a_xlsx(["a"], [{"a": "=cmd"}]))).active
    assert ws is not None
    assert ws.cell(row=2, column=1).value == "'=cmd"


def test_xlsx_control_char_no_crashea() -> None:
    """celda con caracteres de control no lanza IllegalCharacterError; se eliminan."""
    ws = load_workbook(BytesIO(a_xlsx(["a"], [{"a": "x\x00y\x0c"}]))).active
    assert ws is not None
    assert ws.cell(row=2, column=1).value == "xy"


def test_csv_control_char_removido() -> None:
    assert "xy" in _csv_text(["a"], [{"a": "x\x00y"}])
